import csv
from io import StringIO, BytesIO # Added BytesIO
from typing import List, Dict, Any
from datetime import datetime, date
from openpyxl import load_workbook
from src.models.schemas import Transaction

class FileParsingError(Exception):
    """Custom exception for file parsing errors."""
    pass

def parse_csv(file_content: str) -> List[Transaction]:
    """
    Parses CSV content into a list of Transaction objects.
    Assumes CSV has headers and columns like 'Date', 'Description', 'Amount'.
    
    Args:
        file_content: The content of the CSV file as a string.
        
    Returns:
        A list of Transaction objects.
        
    Raises:
        FileParsingError: If required columns are missing or data conversion fails.
    """
    transactions: List[Transaction] = []
    f = StringIO(file_content)
    reader = csv.DictReader(f)

    for row in reader:
        try:
            # Basic mapping, needs to be more robust for varied formats (FR-011)
            # For now, assume 'Date', 'Description', 'Amount'
            transaction_date = datetime.strptime(row['Date'], '%Y-%m-%d').date() # Example format
            description = row['Description']
            amount = float(row['Amount'])
            
            transactions.append(
                Transaction(
                    date=transaction_date,
                    description=description,
                    amount=amount
                )
            )
        except KeyError as e:
            raise FileParsingError(f"Missing expected column in CSV: {e}. Ensure 'Date', 'Description', 'Amount' are present.")
        except ValueError as e:
            raise FileParsingError(f"Data type conversion error in CSV: {e}. Check 'Date' and 'Amount' formats.")
    return transactions

def parse_excel(file_content: bytes) -> List[Transaction]:
    """
    Parses Excel content (bytes) into a list of Transaction objects.
    Assumes first sheet, first row headers, and columns like 'Date', 'Description', 'Amount'.
    
    Args:
        file_content: The content of the Excel file as bytes.
        
    Returns:
        A list of Transaction objects.
        
    Raises:
        FileParsingError: If required columns are missing or data conversion fails.
    """
    transactions: List[Transaction] = []
    
    # Load workbook from bytes
    workbook = load_workbook(filename=BytesIO(file_content)) # Corrected to BytesIO
    sheet = workbook.active
    
    headers = [cell.value for cell in sheet[1]]
    
    for row_idx in range(2, sheet.max_row + 1): # Start from second row for data
        row_data = {}
        for col_idx, header in enumerate(headers):
            row_data[header] = sheet.cell(row=row_idx, column=col_idx + 1).value
        
        try:
            # Handle date parsing from Excel, which can be datetime objects
            if isinstance(row_data['Date'], datetime):
                transaction_date = row_data['Date'].date()
            else:
                transaction_date = datetime.strptime(str(row_data['Date']), '%Y-%m-%d').date()
            description = str(row_data['Description'])
            amount = float(row_data['Amount'])

            transactions.append(
                Transaction(
                    date=transaction_date,
                    description=description,
                    amount=amount
                )
            )
        except KeyError as e:
            raise FileParsingError(f"Missing expected column in Excel: {e}. Ensure 'Date', 'Description', 'Amount' are present.")
        except ValueError as e:
            raise FileParsingError(f"Data type conversion error in Excel: {e}. Check 'Date' and 'Amount' formats.")
        except AttributeError as e:
            raise FileParsingError(f"Date format error in Excel: {e}. Ensure 'Date' column contains valid date formats.")
            
    return transactions
